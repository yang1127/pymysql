# -*-coding:utf-8-*-
# excel文件数据导入数据库
import pymysql
import openpyxl

# 创建连接
conn = pymysql.connect(
    host="127.0.0.1",
    port=3306,
    user="root",
    password="yyy1536520643",
    database="books",
    charset="utf8",
    autocommit=True
)

# 获取游标
cursor = conn.cursor()

# 先创建数据表
cursor.execute('drop table if EXISTS t_game')
sql = '''
    create table t_game(
        rank_id INT ( 10 ) UNSIGNED auto_increment COMMENT 'rank_id', # auto_increment 为主键时自增
        app_id VARCHAR ( 25 ) NOT NULL COMMENT '应用id',
        app_NAME VARCHAR ( 25 ) NOT NULL COMMENT '应用名称',
        class VARCHAR ( 25 ) NOT NULL COMMENT '应用类别',
        Company_NAME VARCHAR ( 25 ) COMMENT '公司名',
        score FLOAT NOT NULL COMMENT '七麦指数',
        PRIMARY KEY ( rank_id )
    )
'''
cursor.execute(sql)

# 创建一个excel工作簿 -- 读取当前路径下文件，2021年12月移动应用增长榜前20.xlsx的数据
excel_workbook = openpyxl.load_workbook("2021年12月移动应用增长榜200名完整版.xlsx")
# 使用默认的sheet表
excel_sheet = excel_workbook.active

# 所有数据的列表
excel_data = []
# 从第二行第一列开始取得数据，不需要第一行表头的数据
# 行、列数在openpyxl库中是从1开始计数的，例如：max_row数据是行数20，处理成第2行到第21行
for row_id in range(2, excel_sheet.max_row+1):
    values = []  # 这一行数据的列表
    for col_id in range(1, excel_sheet.max_column+1):
        values.append(excel_sheet.cell(row_id, col_id).value)
    print(values)
    # 实现2层的列表嵌套 -- 每行values获取的数据，再追加到excel_data中
    excel_data.append(values)
    # print(excel_data)  # 列表嵌套的形式存储
try:
    with conn.cursor() as cursor:
        # cursor.executemany 表示批量插入数据，批处理
        cursor.executemany(
            'insert into t_game'
            '(rank_id, app_id, app_NAME, class, Company_NAME, score)'
            'values'
            '(%s,%s,%s,%s,%s,%s)', excel_data
        )
    conn.commit()
    print("导入成功！")
# 捕获异常
except pymysql.MySQLError as err:
    print(err)  # 如果出现异常，打印错误信息
    print("导入失败！")
finally:
    conn.close()  # 关闭连接，节省资源占用
