# 数据库导出数据为excel文件
import pymysql
import openpyxl  # excel的库

conn = pymysql.connect(
    host="127.0.0.1",
    port=3306,
    user="root",
    password="yyy1536520643",
    database="books",
    autocommit=False,
    charset="utf8mb4"
)

# 从数据库中读取到t_test的sno,sname,sex,age,score并放入excel中
try:
    # 创建游标 cursor = conn.cursor()
    with conn.cursor() as cursor:
        cursor.execute(
            'select sno, sname, sex, age, score from t_test'
        )
        # 创建一个excel工作簿
        excel_workbook = openpyxl.Workbook()
        # 在工作簿中建立一个工作学生成绩表，作为当前要写入内容的表
        # excel_sheet=excel_workbook.create_sheet("学生成绩表")
        # 使用默认的sheet表，不新建表, 切换当前把数据写入此表
        excel_sheet = excel_workbook.active

        count = -1  # 统计文本数据行数
        # openpyxl操作excel，是行和列的索引从1开始，所以要+1，col_id, col_name表示索引和索引内容
        for col_id, col_name in enumerate(['sno', 'sname', 'sex', 'age', 'score']):  # 列索引，读取哪些列的数据
            excel_sheet.cell(1, col_id+1, col_name)  # 往单元格写入数据，第1列表格的表头'sno', 'sname', 'sex', 'age', 'score'写入
            count += 1
        # 写入数据
        # 注意数据量比较少的时候用fetchall()，一行一行的读取
        for row_id, row_emp in enumerate(cursor.fetchall()):  # 获得每一行的数据
            for col_id, col_value in enumerate(row_emp):  # 把每一行的每一列写入，加2是因为原先表头+1的基础上，再加上表头这一行，所以要+1+1=+2
                excel_sheet.cell(row_id+2, col_id+1, col_value)
    excel_workbook.save("读取到的学生成绩.xlsx")
    print(count)
# 捕获异常
except pymysql.MySQLError as err:
    print(err)  # 如果出现异常，打印错误信息
finally:
    conn.close()  # 关闭连接
